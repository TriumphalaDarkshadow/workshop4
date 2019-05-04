from openpyxl import Workbook
from openpyxl import load_workbook
import os

 ##Input soll als Name des Excel-File gespeichert werden
 ##falls Datei existiert (-> pruefen, ob existiert), soll ein weiteres worksheet erstellt werden

if not os.path.exists("Rechnung"):
  os.mkdir("Rechnung")


vorname = input("Geben Sie Ihren Vornamen ein: ")
nachname = input("Geben Sie Ihren Nachnamen ein: ")


sheet_count=1
#lade vorhandenes workbook falls es exestiert
if os.path.exists("Rechnung/"+vorname+nachname+".xlsx"):
  print "workbook exestiert bereits"
  wb=load_workbook("Rechnung/"+vorname+nachname+".xlsx")

   #zähle anzahal der worksheets...gitb warscheinlich auch eine funktion die das macht
  for sheet in wb:
      sheet_count +=1

  print "erstelle worksheet mit nummer" + str(sheet_count)
  
  # erstelle neues worksheet
  ws = wb.create_sheet(str(sheet_count))
  #setze erstelltes worksheet auf als das active
  wb.active = ws
else:
  print "erstelle neues workbook"
  wb = Workbook()  
  ws = wb.active
  


ws.title =str(sheet_count)



stueckpreis_umschlag = 10
stueckpreis_stift = 65
stueckpreis_lineal = 199
stueckpreis_marker = 255

#Kunden Namen einlesen




ws['A1'] = 'Rechnung'
ws['A3'] = 'Vorname'
ws['A4'] = 'Nachname'

ws['B3'] = vorname
ws['B4'] = nachname



#Einlesen der Mengen
print("\n\nGeben Sie die Anzahl der jeweils gekauften Artikel ein.\n")
zahl_umschlag = int(input("Briefumschlaege: "))
zahl_stift = int(input("Bleistifte: "))
zahl_lineal = int(input("Lineale: "))
zahl_marker = int(input("Textmarker: "))

ws['C7'] = zahl_umschlag
ws['C8'] = zahl_stift
ws['C9'] = zahl_lineal
ws['C10'] = zahl_marker

#einkauf = eine Liste von Tupel die die Werte der auszugebenden Tabelle enthlt
einkauf = []
einkauf.append(('Briefumschlag',stueckpreis_umschlag/100, zahl_umschlag,(stueckpreis_umschlag * zahl_umschlag)/100))
einkauf.append(('Bleistift',stueckpreis_stift/100,zahl_stift,(stueckpreis_stift * zahl_stift)/100))
einkauf.append(('Lineal',stueckpreis_lineal/100,zahl_lineal,(stueckpreis_lineal * zahl_lineal)/100))
einkauf.append(('Textmarker',stueckpreis_marker/100,zahl_marker,(stueckpreis_marker * zahl_marker)/100))


#Berechnung des Bruttopreises
brutto_euro = 0
for artikel in einkauf:
    brutto_euro += artikel[3]


#Trennen von Brutto und Netto, ermitteln der Mehrwertsteuer
netto_euro = float("{0:10.2f}".format(brutto_euro/1.19))
mehrwert = float("{0:10.2f}".format(brutto_euro - netto_euro))


#Ausgabe
print("\n\nDer Kassenbon von",vorname,nachname,"\n")
print("Artikel        Stk.-Preis/Euro    Anzahl      Preis/Euro")
print("--------------------------------------------------------")  
 
for artikel in einkauf:
    print("{:<17}".format(artikel[0]), format(artikel[1],'10.2f'), format(artikel[2],'12'), format(artikel[3],'14.2f'))

ws['D7'] = (stueckpreis_umschlag * zahl_umschlag)/100
ws['D8'] = (stueckpreis_stift * zahl_stift)/100
ws['D9'] = (stueckpreis_lineal * zahl_lineal)/100
ws['D10'] = (stueckpreis_marker * zahl_marker)/100


print("\n")
print("                                Zahlbetrag    {0:10.2f}\n".format(brutto_euro))
print("19% MWST{0:10.2f}\nNetto   {1:10.2f} \n\n".format(mehrwert, netto_euro ))

ws['A6'] = 'Artikel'
ws['B6'] = 'Stk.-Preis/Euro'
ws['C6'] = 'Anzahl'
ws['D6'] = 'Preis/Euro'

ws['A7'] = 'Briefumschlag'
ws['A8'] = 'Bleistift'
ws['A9'] = 'Lineal'
ws['A10'] = 'Textmarker'


ws['A14'] = '19% MWST'
ws['D14'] = mehrwert
ws['D15'] = netto_euro
ws['A15'] = 'Netto'
ws['A12'] = 'Zahlbetrag'
zahlbetrag = mehrwert + netto_euro
ws['D12'] = zahlbetrag

ws['B7'] = 0.10
ws['B8'] = 0.65
ws['B9'] = 1.99
ws['B10'] = 2.55

##ws['D7'] = ['=B7*C7']
##ws.write('D7', '=B7*C7')



## < Arbeitsordner>\Rechnung\Rechnung_<Vorname><Nachname>.xlsx
wb.save('Rechnung/'+vorname+nachname+'.xlsx')

